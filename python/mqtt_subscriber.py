import paho.mqtt.client as mqtt
import time
import re
from openpyxl import *
wb = load_workbook("/home/pi/Documents/AS2020/mqtt/test01.xlsx")
ws = wb["Import"]
#global cntr
cntr = 2

MQTT_SERVER = "localhost"
MQTT_PATH = "test_channel"

a2 = 0
buf1 = []
buf2 = []
#cntr = 0
# The callback for when the client receives a CONNACK response from the server.
def on_connect(client, userdata, flags, rc):
    print("Connected with result code " + str(rc))

    # Subscribing in on_connect() means that if we lose the connection and
    # reconnect then subscriptions will be renewed.
    client.subscribe(MQTT_PATH)


# The callback for when a PUBLISH message is received from the server.
def on_message(client, userdata, msg):
    #print(msg.topic + " " + str(msg.payload))
    global cntr
    a1 = msg.payload
    a2 = a1.decode('utf-8')
    print(a2)
    buf1.append(a2)
    buf2 = str(buf1)
    print(buf2)
    temp = re.findall(r'\d+', buf2)
    print(temp)
    res = list(map(int, temp))
    print(res)
    first_i = []
    second_i = []
    third_i = []
    for j in range(0, len(res)):
        if j % 3 == 0:
            first_i.append(res[j])
        elif j % 3 == 1:
            second_i.append(res[j])
        elif j % 3 == 2:
            third_i.append(res[j])

    print(first_i)
    print(second_i)
    print(third_i)
    for k in range(0, len(first_i)):
        wcell1 = ws.cell(cntr, 1)
        wcell1.value = first_i[k]
        wcell1 = ws.cell(cntr, 2)
        wcell1.value = second_i[k]
        wcell1 = ws.cell(cntr, 3)
        wcell1.value = third_i[k]
        cntr = cntr + 1
        wb.save("/home/pi/Documents/AS2020/mqtt/test01.xlsx")
    #connectionSocket.close()
    print("finished")
    #buf1 = []


    # more callbacks, etc


client = mqtt.Client()
client.on_connect = on_connect
client.on_message = on_message

client.connect(MQTT_SERVER, 1883, 60)

# Blocking call that processes network traffic, dispatches callbacks and
# handles reconnecting.
# Other loop*() functions are available that give a threaded interface and a
# manual interface.
client.loop_forever()