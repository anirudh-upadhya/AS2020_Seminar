from socket import *
import re

from openpyxl import *
wb = load_workbook("C:/Users/Anirudh/Documents/SIC/Sem2/Seminar/IOT/backup/WriteTCP.xlsx")
ws = wb["Import"]
cntr = 2
k = 0
x =0
serverPort = 11000
serverSocket = socket(AF_INET, SOCK_STREAM)
serverSocket.bind(('192.168.1.100', serverPort))
serverSocket.listen(50)
print("The Back-Up server is ready to receive")
clientBuf = []
clientData2 = str(0)
while True:
    connectionSocket, addr = serverSocket.accept()
    print(connectionSocket)
    #while True:
    #for i in range(x, x+30):
    while True:
        clientData = connectionSocket.recv(4096).decode()

        if not clientData : break
        elif clientData == 'NP':
            print("Receiving")
            clientData = clientData2
            print(clientData)
        clientBuf.append(clientData)
    #x = i
        print(clientBuf)
        tempBuf = str(clientBuf)
    #print(tempBuf)
        temp = re.findall(r'\d+', tempBuf)
    #print(temp)
        res = list(map(int, temp))
        print(res)
        first_i = []
        second_i = []
        third_i = []
    #odx = res[::2]
    #evx = res[::1]
    #print(odx)
    #print(evx)
        for j in range(0, len(res)):
            if j % 3 == 0:
                first_i.append(res[j])
            elif j % 3 == 1:
                second_i.append(res[j])
            elif j % 3 == 2:
                third_i.append(res[j])

        print(first_i)
        print(second_i)
        print(third_i)
        print(res)
    #for k in range(0, len(first_i)):
        wcell1 = ws.cell(cntr, 1)
        wcell1.value = first_i[k]
        wcell1 = ws.cell(cntr, 2)
        wcell1.value = second_i[k]
        wcell1 = ws.cell(cntr, 3)
        wcell1.value = third_i[k]
        k = k + 1
        cntr = cntr + 1
        wb.save("C:/Users/Anirudh/Documents/SIC/Sem2/Seminar/IOT/backup/WriteTCP.xlsx")
    #connectionSocket.close()
        print("finished")
        clientData2 = clientData
    #clientBuf = 0