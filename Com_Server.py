from socket import *
import re

from openpyxl import *
wb = load_workbook("C:/Users/Anirudh/Documents/SIC/Sem2/Seminar/IOT/WriteTCP.xlsx")
ws = wb["Import"]
cntr = 2
serverPort = 12000
serverSocket = socket(AF_INET, SOCK_STREAM)
serverSocket.bind(('192.168.1.249', serverPort))
serverSocket.listen(50)
print("The server is ready to receive")
clientBuf = []
while True:
    connectionSocket, addr = serverSocket.accept()
    for i in range(0, 30):
        clientData = connectionSocket.recv(32).decode()
        clientBuf.append(clientData)

    print(clientBuf)
    tempBuf = str(clientBuf)
    temp = re.findall(r'\d+', tempBuf)
    res = list(map(int, temp))
    odd_i = []
    even_i = []
    for j in range(0, len(res)):
        if j % 2:
            even_i.append(res[j])
        else:
            odd_i.append(res[j])

    print(odd_i)
    print(even_i)
    print(res)
    for k in range(0, len(odd_i)):
        wcell1 = ws.cell(cntr, 1)
        wcell1.value = odd_i[k]
        wcell1 = ws.cell(cntr, 2)
        wcell1.value = even_i[k]
        cntr = cntr + 1
        wb.save("C:/Users/Anirudh/Documents/SIC/Sem2/Seminar/IOT/WriteTCP.xlsx")
    connectionSocket.close()
